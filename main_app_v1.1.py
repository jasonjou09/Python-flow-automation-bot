# -*- coding: utf-8 -*-
"""
Created on Thu Jul 16 13:55:58 2026

@author: Five-seveN
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import json
import os
import sys
import sqlite3
import pandas as pd
import openpyxl
import pypdf
import pytesseract
from PIL import Image, ImageGrab
from io import BytesIO

# --- 修正 Windows 螢幕縮放導致截圖座標偏移的問題 ---
import ctypes
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
except Exception:
    pass

# --- 剪貼簿控制套件 ---
try:
    import win32clipboard
except ImportError:
    win32clipboard = None

# --- OCR 引擎設定 ---
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

CONFIG_FILE = "app_config.json"
DB_FILE = "tax_billing.db"

class TerminalLogger:
    def __init__(self, text_widget):
        self.text_widget = text_widget
    def write(self, message):
        self.text_widget.insert(tk.END, message)
        self.text_widget.see(tk.END)
    def flush(self): pass

class DatabaseManager:
    def __init__(self, db_name=DB_FILE):
        self.conn = sqlite3.connect(db_name)
        self.cursor = self.conn.cursor()
        self.init_db()

    def init_db(self):
        # 1. 公司基本資訊表 (核心主表)
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_code TEXT UNIQUE, -- 啟用時唯一，停用時為 NULL 以釋出編號
                full_name TEXT NOT NULL,
                short_name TEXT,
                tax_id TEXT,
                base_accounting_fee REAL DEFAULT 0,
                bank_account TEXT,
                is_active INTEGER DEFAULT 1 -- 1:啟用, 0:停用
            )
        ''')
        # 2. 期別帳款明細表 (三維關係表)
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS billing_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_internal_id INTEGER,
                year TEXT,
                month_period TEXT,
                business_tax_amount REAL DEFAULT 0, -- 1. 營業稅 (OCR)
                accounting_fee REAL DEFAULT 0,      -- 2. 記帳費 (歷史快照鎖定)
                note TEXT DEFAULT '',               -- 5. 備註
                is_paid INTEGER DEFAULT 0,
                is_billed INTEGER DEFAULT 0,        -- 已請款狀態 (0: 未請款, 1: 已請款)
                FOREIGN KEY (company_internal_id) REFERENCES companies(id),
                UNIQUE(company_internal_id, year, month_period)
            )
        ''')
        # 3. 動態特別項目明細表
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS billing_special_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                billing_record_id INTEGER,
                item_name TEXT,                     -- 3. 其他特別項目名稱
                item_amount REAL DEFAULT 0,         -- 4. 其他特別項目金額
                FOREIGN KEY (billing_record_id) REFERENCES billing_records(id)
            )
        ''')
        self.conn.commit()
        
        # --- 資料庫結構自動升級 (防呆機制) ---
        try:
            self.cursor.execute("ALTER TABLE billing_records ADD COLUMN is_billed INTEGER DEFAULT 0")
            self.conn.commit()
            print("⚙️ 資料庫成功升級：已加入「已請款」狀態追蹤欄位。")
        except sqlite3.OperationalError:
            # 欄位已存在，略過不處理
            pass

class PeriodSelectionDialog(simpledialog.Dialog):
    """專用的年月份獨立確認彈窗"""
    def __init__(self, parent, title, parent_app):
        self.parent_app = parent_app
        super().__init__(parent, title)
        
    def body(self, master):
        tk.Label(master, text="請明確確認要操作的年月份：", font=("微軟正黑體", 11, "bold")).grid(row=0, columnspan=2, pady=10)
        
        tk.Label(master, text="年份 (民國):").grid(row=1, column=0, sticky="e", padx=5, pady=5)
        self.year_entry = tk.Entry(master, width=12)
        self.year_entry.insert(0, self.parent_app.year_var.get())
        self.year_entry.grid(row=1, column=1, sticky="w", padx=5, pady=5)
        
        tk.Label(master, text="雙月期別:").grid(row=2, column=0, sticky="e", padx=5, pady=5)
        self.period_cb = ttk.Combobox(master, values=self.parent_app.months, state="readonly", width=12)
        self.period_cb.set(self.parent_app.current_month_var.get())
        self.period_cb.grid(row=2, column=1, sticky="w", padx=5, pady=5)
        return self.year_entry
        
    def apply(self):
        self.result = (self.year_entry.get().strip(), self.period_cb.get())

class TaxBillingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("會計報稅請款整合系統 v6.1")
        self.root.geometry("1350x700")
        
        self.db = DatabaseManager()
        self.config = self.load_config()
        self.months = ['01-02月', '03-04月', '05-06月', '07-08月', '09-10月', '11-12月']
        self.month_cols = {'01-02月': 'D', '03-04月': 'E', '05-06月': 'F', '07-08月': 'I', '09-10月': 'J', '11-12月': 'K'}
        
        self.current_month_var = tk.StringVar(value=self.months[0])
        self.selected_record_id = None 
        self.selected_company_id = None
        self.current_record_is_billed = False
        self.current_record_is_paid = False
        
        self.setup_main_layout()
        self.redirect_logging()
        
        print("🚀 系統高度精進版啟動成功！")
        if win32clipboard is None:
            print("⚠️ 警告：未偵測到 pywin32 套件，『一鍵複製為圖片』功能將無法使用。")

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {"year": "115"}

    def save_config(self, *args):
        self.config['year'] = self.year_var.get()
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(self.config, f, ensure_ascii=False, indent=4)

    def setup_main_layout(self):
        # 建立頂層主分頁系統
        self.main_notebook = ttk.Notebook(self.root)
        self.main_notebook.pack(fill=tk.BOTH, expand=True)
        
        self.billing_tab = tk.Frame(self.main_notebook)
        self.company_tab = tk.Frame(self.main_notebook)
        
        self.main_notebook.add(self.billing_tab, text=" 📊 每期帳款與請款管理 ")
        self.main_notebook.add(self.company_tab, text=" 🏢 公司基本資料維護 ")
        
        self.setup_billing_ui()
        self.setup_company_ui()

    def setup_billing_ui(self):
        # === 頂部控制區 ===
        top_frame = tk.Frame(self.billing_tab, padx=10, pady=5)
        top_frame.pack(side=tk.TOP, fill=tk.X)
        
        tk.Label(top_frame, text="設定目前年份:", font=("微軟正黑體", 12)).pack(side=tk.LEFT)
        self.year_var = tk.StringVar(value=self.config.get("year", "115"))
        self.year_var.trace("w", self.save_config)
        tk.Entry(top_frame, textvariable=self.year_var, width=6, font=("微軟正黑體", 12)).pack(side=tk.LEFT, padx=5)
        
        tk.Button(top_frame, text="1. 匯入請款表 (比對底冊)", bg="#9C27B0", fg="white", font=("微軟正黑體", 10, "bold"), command=self.import_billing_sheet).pack(side=tk.LEFT, padx=5)
        tk.Button(top_frame, text="2. 匯入對照表", bg="#2196F3", fg="white", font=("微軟正黑體", 10, "bold"), command=self.import_mapping).pack(side=tk.LEFT, padx=5)
        tk.Button(top_frame, text="3. 讀取 PDF 擷取圖片", bg="#4CAF50", fg="white", font=("微軟正黑體", 10, "bold"), command=self.process_pdfs).pack(side=tk.LEFT, padx=5)
        tk.Button(top_frame, text="4. OCR 辨識金額入庫", bg="#FF5722", fg="white", font=("微軟正黑體", 10, "bold"), command=self.run_ocr).pack(side=tk.LEFT, padx=5)

        # === 中間主工作區 ===
        main_work_frame = tk.Frame(self.billing_tab)
        main_work_frame.pack(fill=tk.BOTH, expand=True)
        
        left_frame = tk.Frame(main_work_frame, padx=10, pady=5)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        right_frame = tk.Frame(main_work_frame, padx=10, pady=5)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # 左側各月份清單
        tk.Label(left_frame, text="各月份公司請款清單 (已依公司編號排序)", font=("微軟正黑體", 12, "bold")).pack(anchor="w")
        self.notebook = ttk.Notebook(left_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
        
        self.tabs = {}
        for m in self.months:
            tab = tk.Frame(self.notebook)
            self.notebook.add(tab, text=m)
            cols = ('billed_status', 'status', 'name', 'bank_account', 'tax_amt', 'accounting_fee', 'special_fee', 'total_amt', 'note', 'id')
            tree = ttk.Treeview(tab, columns=cols, show='tree headings')
            
            tree.heading('#0', text='公司編號')
            tree.column('#0', width=45, anchor='center')
            tree.heading('billed_status', text='請款')
            tree.column('billed_status', width=35, anchor='center')
            tree.heading('status', text='收款')
            tree.column('status', width=35, anchor='center')
            tree.heading('name', text='公司名稱')
            tree.column('name', width=120, anchor='w')
            tree.heading('bank_account', text='對方銀行帳戶')
            tree.column('bank_account', width=150, anchor='w')
            tree.heading('tax_amt', text='營業稅')
            tree.column('tax_amt', width=60, anchor='e')
            tree.heading('accounting_fee', text='記帳費')
            tree.column('accounting_fee', width=60, anchor='e')
            tree.heading('special_fee', text='特別項目')
            tree.column('special_fee', width=60, anchor='e')
            tree.heading('total_amt', text='總金額')
            tree.column('total_amt', width=70, anchor='e')
            tree.heading('note', text='備註')
            tree.column('note', width=60, anchor='w')
            tree.column('id', width=0, stretch=tk.NO)
            
            tree.bind('<ButtonRelease-1>', self.on_tree_select)
            tree.pack(fill=tk.BOTH, expand=True)
            self.tabs[m] = tree

        # 右側請款編輯區
        billing_frame = tk.LabelFrame(right_frame, text="正式請款單編輯區", font=("微軟正黑體", 12))
        billing_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.billing_canvas = tk.Frame(billing_frame, bg="white", padx=10, pady=10)
        self.billing_canvas.pack(fill=tk.BOTH, expand=True)
        tk.Label(self.billing_canvas, text="記帳費請款單", font=("標楷體", 20, "bold"), bg="white").grid(row=0, column=0, columnspan=2, pady=10)
        
        tk.Label(self.billing_canvas, text="公司名稱：", font=("標楷體", 13), bg="white").grid(row=1, column=0, sticky="e", pady=4)
        self.entry_name = tk.Entry(self.billing_canvas, font=("標楷體", 13), width=24, state="readonly")
        self.entry_name.grid(row=1, column=1, sticky="w", pady=4)
        
        tk.Label(self.billing_canvas, text="1. 營業稅：", font=("標楷體", 13), bg="white").grid(row=2, column=0, sticky="e", pady=4)
        self.entry_tax = tk.Entry(self.billing_canvas, font=("標楷體", 13), width=24)
        self.entry_tax.grid(row=2, column=1, sticky="w", pady=4)
        self.entry_tax.bind("<KeyRelease>", lambda e: self.update_live_total())
        
        tk.Label(self.billing_canvas, text="2. 當期記帳費：", font=("標楷體", 13), bg="white").grid(row=3, column=0, sticky="e", pady=4)
        self.entry_accounting_fee = tk.Entry(self.billing_canvas, font=("標楷體", 13), width=24)
        self.entry_accounting_fee.grid(row=3, column=1, sticky="w", pady=4)
        self.entry_accounting_fee.bind("<KeyRelease>", lambda e: self.update_live_total())

        tk.Label(self.billing_canvas, text="5. 備註事項：", font=("標楷體", 13), bg="white").grid(row=4, column=0, sticky="e", pady=4)
        self.entry_note = tk.Entry(self.billing_canvas, font=("標楷體", 13), width=24)
        self.entry_note.grid(row=4, column=1, sticky="w", pady=4)

        # 動態特別繳款項目區
        sp_frame = tk.LabelFrame(self.billing_canvas, text="3 & 4. 動態特別繳款項目項目", bg="white", font=("微軟正黑體", 10))
        sp_frame.grid(row=5, column=0, columnspan=2, pady=5, sticky="we")
        
        self.special_items_tree = ttk.Treeview(sp_frame, columns=('name', 'amount', 'id'), show='headings', height=3)
        self.special_items_tree.heading('name', text='項目名稱')
        self.special_items_tree.heading('amount', text='項目金額')
        self.special_items_tree.column('name', width=120, anchor='w')
        self.special_items_tree.column('amount', width=80, anchor='e')
        self.special_items_tree.column('id', width=0, stretch=tk.NO)
        self.special_items_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        sp_ctrl = tk.Frame(sp_frame, bg="white")
        sp_ctrl.pack(side=tk.RIGHT, fill=tk.Y, padx=5)
        self.entry_sp_name = tk.Entry(sp_ctrl, font=("微軟正黑體", 9), width=12)
        self.entry_sp_name.insert(0, "項目名稱")
        self.entry_sp_name.pack(pady=2)
        self.entry_sp_amount = tk.Entry(sp_ctrl, font=("微軟正黑體", 9), width=12)
        self.entry_sp_amount.insert(0, "金額")
        self.entry_sp_amount.pack(pady=2)
        tk.Button(sp_ctrl, text="➕ 新增", font=("微軟正黑體", 9), bg="#E0E0E0", command=self.add_special_item_click).pack(fill=tk.X, pady=1)
        tk.Button(sp_ctrl, text="❌ 刪除", font=("微軟正黑體", 9), bg="#FFCDD2", command=self.delete_special_item_click).pack(fill=tk.X, pady=1)

        # 最終總和即時回顯
        self.lbl_total_amount = tk.Label(self.billing_canvas, text="最終總款項：0 元", font=("標楷體", 13, "bold"), fg="blue", bg="white")
        self.lbl_total_amount.grid(row=6, column=0, columnspan=2, pady=5)

        btn_frame = tk.Frame(self.billing_canvas, bg="white")
        btn_frame.grid(row=7, column=0, columnspan=2, pady=5)
        
        tk.Button(btn_frame, text="💾 儲存修改", font=("微軟正黑體", 11), bg="#E0E0E0", command=self.save_record_edit).pack(side=tk.LEFT, padx=5)
        
        # 新增的已請款按鈕，放置於付款確認的左邊
        self.btn_mark_billed = tk.Button(btn_frame, text="📨 確認已請款", font=("微軟正黑體", 11), bg="#009688", fg="white", command=self.toggle_billed_status)
        self.btn_mark_billed.pack(side=tk.LEFT, padx=5)
        
        self.btn_mark_paid = tk.Button(btn_frame, text="✅ 確認已付款 (回寫總表)", font=("微軟正黑體", 11), bg="#8BC34A", command=self.toggle_paid_status)
        self.btn_mark_paid.pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="📄 產生請款單", font=("微軟正黑體", 11), bg="#FFC107", command=self.generate_billing_excel).pack(side=tk.LEFT, padx=5)

        # 系統終端機
        term_frame = tk.LabelFrame(right_frame, text="系統狀態回報區", font=("微軟正黑體", 11))
        term_frame.pack(fill=tk.BOTH, expand=True)
        self.terminal = tk.Text(term_frame, height=10, bg="#1E1E1E", fg="#00FF00", font=("Consolas", 10))
        self.terminal.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    def setup_company_ui(self):
        # === 公司基本資料維護介面 ===
        left_pane = tk.Frame(self.company_tab, padx=10, pady=5)
        left_pane.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        right_pane = tk.Frame(self.company_tab, padx=10, pady=5, width=350)
        right_pane.pack(side=tk.RIGHT, fill=tk.BOTH, expand=False)
        
        # 左半部：分組列表
        tk.Label(left_pane, text="🟢 啟用中公司基本資料 (依編號排序)", font=("微軟正黑體", 11, "bold"), fg="green").pack(anchor="w", pady=(5,0))
        cols = ('code', 'full_name', 'short_name', 'tax_id', 'fee', 'bank', 'id')
        self.active_company_tree = ttk.Treeview(left_pane, columns=cols, show='tree headings', height=14)
        self.active_company_tree.heading('#0', text='公司編號')
        self.active_company_tree.column('#0', width=80, anchor='center')
        self.active_company_tree.heading('full_name', text='公司完整名稱')
        self.active_company_tree.column('full_name', width=180, anchor='w')
        self.active_company_tree.heading('short_name', text='簡稱')
        self.active_company_tree.column('short_name', width=100, anchor='w')
        self.active_company_tree.heading('tax_id', text='統一編號')
        self.active_company_tree.column('tax_id', width=90, anchor='center')
        self.active_company_tree.heading('fee', text='月記帳費')
        self.active_company_tree.column('fee', width=80, anchor='e')
        self.active_company_tree.heading('bank', text='對方銀行帳戶')
        self.active_company_tree.column('bank', width=180, anchor='w')
        self.active_company_tree.column('id', width=0, stretch=tk.NO)
        self.active_company_tree.bind('<ButtonRelease-1>', self.on_company_tree_select)
        self.active_company_tree.pack(fill=tk.BOTH, expand=True, pady=5)
        
        tk.Label(left_pane, text="🔴 已停用/歷史公司封存清單 (編號已釋出)", font=("微軟正黑體", 11, "bold"), fg="red").pack(anchor="w", pady=(10,0))
        self.inactive_company_tree = ttk.Treeview(left_pane, columns=cols, show='headings', height=8)
        self.inactive_company_tree.heading('full_name', text='公司完整名稱')
        self.inactive_company_tree.column('full_name', width=200, anchor='w')
        self.inactive_company_tree.heading('short_name', text='簡稱')
        self.inactive_company_tree.column('short_name', width=100, anchor='w')
        self.inactive_company_tree.heading('tax_id', text='統一編號')
        self.inactive_company_tree.column('tax_id', width=100, anchor='center')
        self.inactive_company_tree.heading('fee', text='月記帳費')
        self.inactive_company_tree.column('fee', width=80, anchor='e')
        self.inactive_company_tree.heading('bank', text='對方銀行帳戶')
        self.inactive_company_tree.column('bank', width=200, anchor='w')
        self.inactive_company_tree.column('id', width=0, stretch=tk.NO)
        self.inactive_company_tree.bind('<ButtonRelease-1>', self.on_company_tree_select)
        self.inactive_company_tree.pack(fill=tk.BOTH, expand=True, pady=5)

        # 右半部：編輯面板
        edit_frame = tk.LabelFrame(right_pane, text="公司主檔維護面板", font=("微軟正黑體", 11), padx=10, pady=10)
        edit_frame.pack(fill=tk.BOTH, expand=True)
        
        self.comp_code_var = tk.StringVar()
        self.comp_full_name_var = tk.StringVar()
        self.comp_short_name_var = tk.StringVar()
        self.comp_tax_id_var = tk.StringVar()
        self.comp_base_fee_var = tk.StringVar()
        self.comp_bank_account_var = tk.StringVar()
        
        fields = [
            ("公司編號(新客填/舊客免):", self.comp_code_var),
            ("公司完整名稱:", self.comp_full_name_var),
            ("公司簡約簡稱:", self.comp_short_name_var),
            ("統一編號(OCR核心):", self.comp_tax_id_var),
            ("標準月記帳費:", self.comp_base_fee_var),
            ("對方銀行帳戶資訊:", self.comp_bank_account_var)
        ]
        for i, (lbl, var) in enumerate(fields):
            tk.Label(edit_frame, text=lbl, font=("微軟正黑體", 10)).pack(anchor="w", pady=(5,0))
            tk.Entry(edit_frame, textvariable=var, font=("微軟正黑體", 11), width=30).pack(fill=tk.X, pady=2)
            
        tk.Button(edit_frame, text="➕ 儲存 / 新增客戶資料", font=("微軟正黑體", 10, "bold"), bg="#E0E0E0", command=self.save_company_info).pack(fill=tk.X, pady=(20, 5))
        tk.Button(edit_frame, text="⛔ 設為停用 (立刻釋出編號)", font=("微軟正黑體", 10), bg="#FFCDD2", fg="darkred", command=self.deactivate_company).pack(fill=tk.X, pady=5)
        tk.Button(edit_frame, text="♻️ 恢復啟用 (重新指定編號)", font=("微軟正黑體", 10), bg="#C8E6C9", fg="darkgreen", command=self.reactivate_company).pack(fill=tk.X, pady=5)
        tk.Button(edit_frame, text="🧹 清空輸入框", font=("微軟正黑體", 10), bg="#F5F5F5", command=self.clear_company_entries).pack(fill=tk.X, pady=5)

        self.refresh_company_tab()

    def redirect_logging(self):
        sys.stdout = TerminalLogger(self.terminal)
        sys.stderr = TerminalLogger(self.terminal)

    def on_tab_changed(self, event):
        selected_tab = self.notebook.tab(self.notebook.select(), "text")
        self.current_month_var.set(selected_tab)
        self.refresh_treeview(selected_tab)

    def check_and_initialize_billing(self, year, month_period):
        """核心自動觸發機制：換年或換期時自動建立快照底冊"""
        self.db.cursor.execute("SELECT id, base_accounting_fee FROM companies WHERE is_active = 1")
        active_comps = self.db.cursor.fetchall()
        
        for c_id, base_fee in active_comps:
            self.db.cursor.execute('''
                SELECT id FROM billing_records 
                WHERE company_internal_id = ? AND year = ? AND month_period = ?
            ''', (c_id, year, month_period))
            if not self.db.cursor.fetchone():
                # 歷史快照鎖定：每期費用 = 月記帳費 * 2
                period_fee = (base_fee * 2) if base_fee else 0.0
                self.db.cursor.execute('''
                    INSERT INTO billing_records (company_internal_id, year, month_period, business_tax_amount, accounting_fee, note, is_paid, is_billed)
                    VALUES (?, ?, ?, 0.0, ?, '', 0, 0)
                ''', (c_id, year, month_period, period_fee))
        self.db.conn.commit()

    # ================= 1. 匯入請款表 (建立/比對底冊) =================
    def import_billing_sheet(self):
        year = self.year_var.get()
        if not messagebox.askyesno("底冊建立確認", f"是否先根據現有啟用中客戶，為【{year}年】全年度快速初始化空的底冊？"):
            return

        print(f"\n正在為【{year}年】所有期別動態初始化基本底冊...")
        for m in self.months:
            self.check_and_initialize_billing(year, m)
        print("✅ 全年度基本底冊已自動就緒！")

        if messagebox.askyesno("加選請款表", "您是否要額外選取『請款表 Excel』來比對排除獨立手動維護項目？"):
            file_path = filedialog.askopenfilename(title="選擇 請款表 Excel 檔案", filetypes=[("Excel files", "*.xlsx *.xls")])
            if not file_path: 
                self.refresh_treeview(self.current_month_var.get())
                return
            
            try:
                print("正在掃描 Excel 各工作表，進行公司名稱比對...")
                all_sheets = pd.read_excel(file_path, sheet_name=None)
                count = 0
                skipped = 0
                
                for sheet_name, df in all_sheets.items():
                    sheet_data = df.to_numpy()
                    if sheet_data.shape[0] == 0 or sheet_data.shape[1] < 2: continue
                    anchor_val = sheet_data[0, 0]
                    for r in range(len(sheet_data)):
                        if sheet_data[r, 0] == anchor_val and not pd.isna(sheet_data[r, 1]):
                            company_name = str(sheet_data[r, 1]).replace(" ", "").strip()
                            if company_name:
                                # 只搜尋目前啟用中的公司
                                self.db.cursor.execute("SELECT id FROM companies WHERE (full_name = ? OR short_name = ?) AND is_active = 1", (company_name, company_name))
                                if self.db.cursor.fetchone():
                                    count += 1
                                else:
                                    skipped += 1
                                    print(f"⚠️ 項目 '{company_name}' (位於分頁:{sheet_name}) 在基本資料中無對應啟用客戶，已自動跳過此獨立手動項。")
                print(f"✅ 比對核對完畢！成功映射 {count} 間已知客戶，優雅跳過 {skipped} 個特殊/獨立維護項目。")
            except Exception as e:
                print(f"❌ 讀取請款表發生錯誤：{str(e)}")
                
        self.refresh_treeview(self.current_month_var.get())

    # ================= 2. 匯入對照表 (五欄位排版) =================
    def import_mapping(self):
        file_path = filedialog.askopenfilename(title="選擇對照表檔案", filetypes=[("Excel/CSV files", "*.xlsx *.xls *.csv")])
        if not file_path: return
        try:
            # 【精進優化】強制指定 dtype=str，確保所有欄位都以純文字讀取，避免開頭的 0 被自動轉成數字而吃掉
            df = pd.read_csv(file_path, dtype=str) if file_path.endswith('.csv') else pd.read_excel(file_path, dtype=str)
            count_new = 0
            count_update = 0
            
            print("\n開始解析對照表五欄位數據...")
            for index, row in df.iterrows():
                if len(row) < 5: continue
                code = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ""
                name = str(row.iloc[1]).strip() if not pd.isna(row.iloc[1]) else ""
                tax_id = str(row.iloc[2]).strip() if not pd.isna(row.iloc[2]) else ""
                fee = row.iloc[3]
                bank_acc = str(row.iloc[4]).strip() if not pd.isna(row.iloc[4]) else ""
                
                if not code or code == 'nan' or code == "": continue
                if name == 'nan': name = ""
                if tax_id == 'nan': tax_id = ""
                if bank_acc == 'nan': bank_acc = ""
                
                try: fee = float(fee) if not pd.isna(fee) and str(fee) != 'nan' else 0.0
                except: fee = 0.0
                
                # 【精進優化】台灣統一編號固定為 8 位數，若外部來源 Excel 檔案已被啃掉 0，在此進行防呆補零
                if tax_id and tax_id.isdigit() and len(tax_id) < 8:
                    tax_id = tax_id.zfill(8)
                
                # 檢查同編號且啟用中的公司
                self.db.cursor.execute("SELECT id FROM companies WHERE company_code = ? AND is_active = 1", (code,))
                res = self.db.cursor.fetchone()
                
                if res:
                    self.db.cursor.execute('''
                        UPDATE companies SET full_name = ?, short_name = ?, tax_id = ?, base_accounting_fee = ?, bank_account = ?
                        WHERE id = ?
                    ''', (name, name, tax_id, fee, bank_acc, res[0]))
                    count_update += 1
                else:
                    self.db.cursor.execute('''
                        INSERT INTO companies (company_code, full_name, short_name, tax_id, base_accounting_fee, bank_account, is_active)
                        VALUES (?, ?, ?, ?, ?, ?, 1)
                    ''', (code, name, name, tax_id, fee, bank_acc))
                    count_new += 1
                    
            self.db.conn.commit()
            print(f"✅ 對照表匯入成功！全新導入 {count_new} 筆，更版升級 {count_update} 筆。")
            self.refresh_company_tab()
            self.refresh_treeview(self.current_month_var.get())
        except Exception as e:
            print(f"❌ 對照表匯入失敗：{str(e)}")

    # ================= 3 & 4. PDF 擷取與 OCR 金額入庫 =================
    def process_pdfs(self):
        dlg = PeriodSelectionDialog(self.root, "確認 PDF 產出期別", self)
        if not dlg.result: return
        year, month = dlg.result
        
        files = filedialog.askopenfilenames(title="選擇國稅局 PDF 檔", filetypes=[("PDF files", "*.pdf")])
        if not files: return
        
        target_folder = f"{year}年_{month}_產出圖片"
        os.makedirs(target_folder, exist_ok=True)

        print(f"\n開始擷取【{year}年 {month}】的 PDF 圖片...")
        success = 0
        for file in files:
            try:
                reader = pypdf.PdfReader(file)
                base_name = os.path.basename(file)
                tax_id = base_name[6:14] 
                with open(os.path.join(target_folder, f"{tax_id}.png"), "wb") as fp:
                    fp.write(reader.pages[0].images[0].data)
                success += 1
            except Exception as e:
                print(f"⚠️ 處理 '{os.path.basename(file)}' 失敗: {str(e)}")
        print(f"✅ 完成！共擷取 {success} 張統編圖片至「{target_folder}」。\n")

    def run_ocr(self):
            dlg = PeriodSelectionDialog(self.root, "確認 OCR 寫入期別", self)
            if not dlg.result: return
            year, month = dlg.result
            
            folder = filedialog.askdirectory(title="選擇存放統編圖片的資料夾")
            if not folder: return
            
            print(f"\n🔍 開始辨識【{year}年 {month}】圖片金額並自動對應...")
            png_files = [f for f in os.listdir(folder) if f.endswith('.png')]
            count = 0
            skipped = 0 # 紀錄因為已經有數值而跳過的數量
            
            for filename in png_files:
                tax_id = filename[0:8]
                try:
                    image = Image.open(os.path.join(folder, filename)).convert('L')
                    crpimg = image.crop((3060, 730, 3220, 780))
                    ocr_text = pytesseract.image_to_string(crpimg).replace("\n", "").replace(",", "").strip()
                    
                    # 強制剔除所有非數字元，只保留純數字，完美避開雜訊誤判
                    clean_text = ''.join(filter(str.isdigit, ocr_text))
                    amount = float(clean_text) if clean_text else 0.0
                    
                    # 根據統編搜尋啟用中的客戶內部 ID，順便撈出完整名稱以利彈窗提示
                    self.db.cursor.execute("SELECT id, full_name FROM companies WHERE tax_id = ? AND is_active = 1", (tax_id,))
                    c_res = self.db.cursor.fetchone()
                    
                    if c_res:
                        c_id, c_name = c_res
                        self.check_and_initialize_billing(year, month)
                        
                        # 【新增：資料防覆蓋機制】先檢查該期該公司的營業稅是否已經有大於 0 的數值
                        self.db.cursor.execute('''
                            SELECT business_tax_amount FROM billing_records 
                            WHERE year = ? AND month_period = ? AND company_internal_id = ?
                        ''', (year, month, c_id))
                        existing_tax_res = self.db.cursor.fetchone()
                        
                        if existing_tax_res and existing_tax_res[0] > 0.0:
                            # 已經有數值，跳出警告並略過 UPDATE
                            print(f"⚠️ 跳過覆蓋：客戶 [{c_name}] 已有營業稅紀錄 ({existing_tax_res[0]:.0f} 元)。")
                            skipped += 1
                        else:
                            # 確定是空的 (或為 0.0)，執行正常寫入
                            self.db.cursor.execute('''
                                UPDATE billing_records SET business_tax_amount = ? 
                                WHERE year = ? AND month_period = ? AND company_internal_id = ?
                            ''', (amount, year, month, c_id))
                            self.db.conn.commit()
                            count += 1
                    else:
                        print(f"⚠️ 找不到統編 {tax_id} 的啟用中客戶基本資料，已自動跳過。")
                except Exception as e:
                    print(f"❌ '{filename}' 辨識失敗: {str(e)}")
                    
            print(f"✅ OCR 完工，已精準填入 {count} 筆營業稅格子中。（觸發保護機制共跳過 {skipped} 筆）")
            
            # 完成後，強制切換 UI 畫面到剛剛辨識的那個月份
            if month in self.months:
                self.notebook.select(self.months.index(month))
                self.current_month_var.set(month)
            self.refresh_treeview(month)

    # ================= 互動編輯與動態項目邏輯 =================
    def on_tree_select(self, event):
        tree = event.widget
        selected = tree.selection()
        if not selected or not tree.item(selected[0])['values']: return
        
        values = tree.item(selected[0])['values']
        self.selected_record_id = values[9] # 新增請款狀態欄位後，ID 的位置往後移至索引 9
        
        # 在 SQL 中多撈取 r.is_paid 與 r.is_billed 狀態
        self.db.cursor.execute('''
            SELECT c.company_code, c.full_name, r.business_tax_amount, r.accounting_fee, r.note, r.is_paid, r.is_billed
            FROM billing_records r JOIN companies c ON r.company_internal_id = c.id WHERE r.id = ?
        ''', (self.selected_record_id,))
        res = self.db.cursor.fetchone()
        if not res: return
        code, name, tax, fee, note, is_paid, is_billed = res
        
        self.entry_name.config(state="normal")
        self.entry_name.delete(0, tk.END)
        self.entry_name.insert(0, f"[{code}] {name}")
        self.entry_name.config(state="readonly")
        
        self.entry_tax.delete(0, tk.END)
        # 精進點：去除小數點 .0。但注意「不要加上千分位逗號」，以防未來 float() 轉型報錯
        self.entry_tax.insert(0, str(int(round(tax))) if tax is not None else "0")
        
        self.entry_accounting_fee.delete(0, tk.END)
        # 精進點：去除小數點 .0。同樣不加逗號
        self.entry_accounting_fee.insert(0, str(int(round(fee))) if fee is not None else "0")
        
        self.entry_note.delete(0, tk.END)
        self.entry_note.insert(0, note if note else "")

        # 根據請款狀態，動態改變按鈕文字、顏色
        self.current_record_is_billed = bool(is_billed)
        if self.current_record_is_billed:
            self.btn_mark_billed.config(text="⏪ 取消請款", bg="#FF9800") # 橘色
        else:
            self.btn_mark_billed.config(text="📨 確認已請款", bg="#009688") # 青綠色

        # 根據付款狀態，動態改變按鈕文字、顏色與內部狀態標記
        self.current_record_is_paid = bool(is_paid)
        if self.current_record_is_paid:
            self.btn_mark_paid.config(text="⏪ 取消確認 (清除總表)", bg="#FF9800") # 變成醒目的橘色
        else:
            self.btn_mark_paid.config(text="✅ 確認已付款 (回寫總表)", bg="#8BC34A") # 恢復原本的綠色
        
        self.load_special_items_to_list()
        self.update_live_total()

    def load_special_items_to_list(self):
        for item in self.special_items_tree.get_children():
            self.special_items_tree.delete(item)
        if not self.selected_record_id: return
        self.db.cursor.execute("SELECT item_name, item_amount, id FROM billing_special_items WHERE billing_record_id = ?", (self.selected_record_id,))
        for n, a, sid in self.db.cursor.fetchall():
            # 精進點：特別繳款項目金額也去除小數點
            self.special_items_tree.insert('', tk.END, values=(n, int(round(a)) if a is not None else 0, sid))

    def update_live_total(self):
        try: tax = float(self.entry_tax.get().strip())
        except: tax = 0.0
        try: fee = float(self.entry_accounting_fee.get().strip())
        except: fee = 0.0
        
        sp_sum = 0.0
        for item in self.special_items_tree.get_children():
            vals = self.special_items_tree.item(item)['values']
            try: sp_sum += float(vals[1])
            except: pass
        
        final_total = tax + fee + sp_sum
        # 精進點：最終即時回顯總金額完全「去小數點 + 格式化為千分位 comma 逗號」顯示，體驗完美
        self.lbl_total_amount.config(text=f"最終總款項：{int(round(final_total)):,} 元 (營業稅:{int(round(tax)):,} + 記帳:{int(round(fee)):,} + 特別項目:{int(round(sp_sum)):,})")

    def add_special_item_click(self):
        if not self.selected_record_id: return
        name = self.entry_sp_name.get().strip()
        amt_str = self.entry_sp_amount.get().strip()
        if not name or name == "項目名稱": return
        try: amt = float(amt_str)
        except: return
        
        self.db.cursor.execute("INSERT INTO billing_special_items (billing_record_id, item_name, item_amount) VALUES (?, ?, ?)", (self.selected_record_id, name, amt))
        self.db.conn.commit()
        self.entry_sp_name.delete(0, tk.END)
        self.entry_sp_amount.delete(0, tk.END)
        self.load_special_items_to_list()
        self.update_live_total()

    def delete_special_item_click(self):
        sel = self.special_items_tree.selection()
        if not sel: return
        sid = self.special_items_tree.item(sel[0])['values'][2]
        self.db.cursor.execute("DELETE FROM billing_special_items WHERE id = ?", (sid,))
        self.db.conn.commit()
        self.load_special_items_to_list()
        self.update_live_total()

    def save_record_edit(self):
        if not self.selected_record_id: return
        try: tax = float(self.entry_tax.get().strip())
        except: tax = 0.0
        try: fee = float(self.entry_accounting_fee.get().strip())
        except: fee = 0.0
        note = self.entry_note.get().strip()
        
        self.db.cursor.execute('UPDATE billing_records SET business_tax_amount = ?, accounting_fee = ?, note = ? WHERE id = ?', (tax, fee, note, self.selected_record_id))
        self.db.conn.commit()
        print("💾 金額及備註主數據修改成功。")
        self.refresh_treeview(self.current_month_var.get())

    # ================= 📨 人工確認已請款 (僅資料庫狀態維護) =================
    def toggle_billed_status(self):
        if not self.selected_record_id: return
        month = self.current_month_var.get()
        
        self.db.cursor.execute('''
            SELECT c.company_code, c.full_name FROM billing_records r 
            JOIN companies c ON r.company_internal_id = c.id WHERE r.id = ?
        ''', (self.selected_record_id,))
        res = self.db.cursor.fetchone()
        if not res or not res[0]:
            messagebox.showerror("錯誤", "該客戶沒有合法的啟用中公司編號，無法操作請款。")
            return
        comp_code, comp_name = res
        
        if self.current_record_is_billed:
            # 【取消請款邏輯】
            self.db.cursor.execute("UPDATE billing_records SET is_billed = 0 WHERE id = ?", (self.selected_record_id,))
            self.db.conn.commit()
            print(f"⏪ 已取消請款狀態：將 [{comp_code}]{comp_name} 於 {month} 的記錄標記為「未請款」。")
            self.current_record_is_billed = False
            self.btn_mark_billed.config(text="📨 確認已請款", bg="#009688")
        else:
            # 【確認請款邏輯】
            self.db.cursor.execute("UPDATE billing_records SET is_billed = 1 WHERE id = ?", (self.selected_record_id,))
            self.db.conn.commit()
            print(f"📨 已請款：已將 [{comp_code}]{comp_name} 於 {month} 的狀態標記為「已請款」。")
            self.current_record_is_billed = True
            self.btn_mark_billed.config(text="⏪ 取消請款", bg="#FF9800")
            
        self.refresh_treeview(month)

    # ================= ✅ 確認已付款 (僅回寫記帳費+特別項目) =================
    def toggle_paid_status(self):
        if not self.selected_record_id: return
        year = self.year_var.get()
        month = self.current_month_var.get()
        
        self.db.cursor.execute('''
            SELECT c.company_code, c.full_name, r.accounting_fee FROM billing_records r 
            JOIN companies c ON r.company_internal_id = c.id WHERE r.id = ?
        ''', (self.selected_record_id,))
        res = self.db.cursor.fetchone()
        if not res or not res[0]:
            messagebox.showerror("錯誤", "該客戶沒有合法的啟用中公司編號，無法操作總表。")
            return
        comp_code, comp_name, act_fee = res
        
        # 排除營業稅：金額 = 當期記帳費 + 所有特別項目加總
        self.db.cursor.execute("SELECT SUM(item_amount) FROM billing_special_items WHERE billing_record_id = ?", (self.selected_record_id,))
        sp_res = self.db.cursor.fetchone()
        sp_sum = sp_res[0] if sp_res[0] else 0.0
        excel_amount = act_fee + sp_sum
        
        master_file = f"{year}年01-12月客戶帳單明細.xlsx"
        try:
            wb = openpyxl.load_workbook(master_file)
            sheet_name = f"{year}年"
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"明細總表找不到名為 '{sheet_name}' 的分頁。")
            
            ws = wb[sheet_name]
            target_col = self.month_cols.get(month)
            row_found = False
            
            # 精準定位總表列數
            for row in range(2, ws.max_row + 1):
                cell_val = str(ws[f'A{row}'].value).replace(" ", "").strip()
                if cell_val == comp_code:
                    if self.current_record_is_paid:
                        # 【取消付款邏輯】清空 Excel 格子
                        ws[f'{target_col}{row}'].value = None
                    else:
                        # 【確認付款邏輯】寫入總金額
                        ws[f'{target_col}{row}'].value = float(excel_amount)
                    row_found = True
                    break
                    
            if not row_found:
                print(f"⚠️ 警告：在總表 A 欄中找不到客戶編號 '{comp_code}'，無法同步 Excel。")
            else:
                wb.save(master_file)
                
                # 更新資料庫狀態與介面按鈕
                if self.current_record_is_paid:
                    print(f"⏪ 已取消確認：將 [{comp_code}]{comp_name} 於 {month} 總表 {target_col} 欄的紀錄抹除。")
                    self.db.cursor.execute("UPDATE billing_records SET is_paid = 0 WHERE id = ?", (self.selected_record_id,))
                    self.current_record_is_paid = False
                    self.btn_mark_paid.config(text="✅ 確認已付款 (回寫總表)", bg="#8BC34A")
                else:
                    print(f"✅ 收款成功：已將 [{comp_code}]{comp_name} 的帳款(不含營業稅) {excel_amount} 元寫入總表 {month} 的 {target_col} 欄。")
                    self.db.cursor.execute("UPDATE billing_records SET is_paid = 1 WHERE id = ?", (self.selected_record_id,))
                    self.current_record_is_paid = True
                    self.btn_mark_paid.config(text="⏪ 取消確認 (清除總表)", bg="#FF9800")
                
                self.db.conn.commit()
                self.refresh_treeview(month)
                
        except FileNotFoundError:
            print(f"❌ 錯誤：找不到總表 '{master_file}'，請確認檔案放置於同目錄下。")
        except PermissionError:
            messagebox.showerror("檔案被鎖定", f"無法修改「{master_file}」！\n請確認總表是否正在 Excel 中開啟，請先關閉視窗後再試一次。")
        except Exception as e:
            print(f"❌ 操作 Excel 總表發生未知錯誤: {str(e)}")

    # ================= 公司基本資料 CRUD 管理面板邏輯 =================
    def on_company_tree_select(self, event):
        tree = event.widget
        sel = tree.selection()
        if not sel: return
        vals = tree.item(sel[0])['values']
        self.selected_company_id = vals[6]
        
        self.comp_code_var.set(str(tree.item(sel[0])['text']) if tree.item(sel[0])['text'] else "")
        self.comp_full_name_var.set(vals[1])
        self.comp_short_name_var.set(vals[2])
        self.comp_tax_id_var.set(vals[3])
        self.comp_base_fee_var.set(str(vals[4]))
        self.comp_bank_account_var.set(vals[5])

    def save_company_info(self):
            code = self.comp_code_var.get().strip()
            full_name = self.comp_full_name_var.get().strip()
            short_name = self.comp_short_name_var.get().strip()
            tax_id = self.comp_tax_id_var.get().strip()
            fee_str = self.comp_base_fee_var.get().strip()
            bank_acc = self.comp_bank_account_var.get().strip()
            
            if not full_name: return
            try: fee = float(fee_str)
            except: fee = 0.0
            
            # 【精進優化】手動介面建檔或修改時，也自動實施 8 位數統編防呆補零
            if tax_id and tax_id.isdigit() and len(tax_id) < 8:
                tax_id = tax_id.zfill(8)
                
            if getattr(self, 'selected_company_id', None):
                if code:
                    self.db.cursor.execute("SELECT id FROM companies WHERE company_code = ? AND id != ? AND is_active = 1", (code, self.selected_company_id))
                    if self.db.cursor.fetchone():
                        messagebox.showerror("錯誤", f"公司編號 '{code}' 目前已被其他啟用中的客戶佔用！")
                        return
                self.db.cursor.execute('''
                    UPDATE companies SET company_code = ?, full_name = ?, short_name = ?, tax_id = ?, base_accounting_fee = ?, bank_account = ?
                    WHERE id = ?
                ''', (code if code else None, full_name, short_name, tax_id, fee, bank_acc, self.selected_company_id))
                print(f"💾 客戶資料更新完成：{full_name}")
    
                # 智慧連動更新：保護歷史紀錄的同時，允許連動修改當前未付款的帳單
                current_year = self.year_var.get()
                if messagebox.askyesno("同步更新確認", f"您修改了客戶基本資料/記帳費。\n是否要將新的記帳費 ({fee*2} 元/期) 同步套用到【{current_year}年】所有『尚未付款』的請款單中？\n\n(已付款的歷史帳單不會受影響)"):
                    period_fee = fee * 2
                    self.db.cursor.execute('''
                        UPDATE billing_records 
                        SET accounting_fee = ? 
                        WHERE company_internal_id = ? AND year = ? AND is_paid = 0
                    ''', (period_fee, self.selected_company_id, current_year))
                    self.db.conn.commit()
                    print(f"✅ 已同步更新 {full_name} 今年度未付款期別的記帳費！")
    
            else:
                if code:
                    self.db.cursor.execute("SELECT id FROM companies WHERE company_code = ? AND is_active = 1", (code,))
                    if self.db.cursor.fetchone():
                        messagebox.showerror("錯誤", f"公司編號 '{code}' 已存在！")
                        return
                self.db.cursor.execute('''
                    INSERT INTO companies (company_code, full_name, short_name, tax_id, base_accounting_fee, bank_account, is_active)
                    VALUES (?, ?, ?, ?, ?, ?, 1)
                ''', (code if code else None, full_name, short_name, tax_id, fee, bank_acc))
                print(f"➕ 全新客戶建檔成功：{full_name}")
                
            self.db.conn.commit()
            self.clear_company_entries()
            self.refresh_company_tab()
            self.refresh_treeview(self.current_month_var.get())

    def deactivate_company(self):
        if not getattr(self, 'selected_company_id', None): return
        if not messagebox.askyesno("解約/停用確認", "確定將此客戶改為停用狀態？\n這將完全釋出其公司編號給新客遞補，但歷史帳單絕對完好保存。"): return
        self.db.cursor.execute("UPDATE companies SET is_active = 0, company_code = NULL WHERE id = ?", (self.selected_company_id,))
        self.db.conn.commit()
        print("⛔ 客戶停用成功，編號已安全釋出。")
        self.clear_company_entries()
        self.refresh_company_tab()
        self.refresh_treeview(self.current_month_var.get())

    def reactivate_company(self):
        if not getattr(self, 'selected_company_id', None): return
        new_code = simpledialog.askstring("恢復啟用", "請輸入指定分配給該客戶的全新公司編號：")
        if not new_code: return
        new_code = new_code.strip()
        
        self.db.cursor.execute("SELECT id FROM companies WHERE company_code = ? AND is_active = 1", (new_code,))
        if self.db.cursor.fetchone():
            messagebox.showerror("錯誤", f"指定編號 '{new_code}' 正被其他客戶使用，無法遞補。")
            return
            
        self.db.cursor.execute("UPDATE companies SET is_active = 1, company_code = ? WHERE id = ?", (new_code, self.selected_company_id))
        self.db.conn.commit()
        print(f"♻️ 客戶已順利重新啟用，獲配代碼：{new_code}")
        self.clear_company_entries()
        self.refresh_company_tab()
        self.refresh_treeview(self.current_month_var.get())

    def clear_company_entries(self):
        self.selected_company_id = None
        self.comp_code_var.set("")
        self.comp_full_name_var.set("")
        self.comp_short_name_var.set("")
        self.comp_tax_id_var.set("")
        self.comp_base_fee_var.set("")
        self.comp_bank_account_var.set("")

    def refresh_company_tab(self):
        for item in self.active_company_tree.get_children(): self.active_company_tree.delete(item)
        for item in self.inactive_company_tree.get_children(): self.inactive_company_tree.delete(item)
        
        # 啟用中公司依編號由小到大排序
        self.db.cursor.execute('SELECT company_code, full_name, short_name, tax_id, base_accounting_fee, bank_account, id FROM companies WHERE is_active=1 ORDER BY company_code ASC')
        for code, f, s, t, fee, b, cid in self.db.cursor.fetchall():
            # 精進點：月記帳費去除小數點 .0，且不含逗號避免未來點選讀取回編輯框時發生 float() 轉型錯誤
            self.active_company_tree.insert('', tk.END, text=code, values=(code, f, s, t, int(round(fee)) if fee else 0, b, cid))
            
        # 停用公司依完整名稱排序
        self.db.cursor.execute('SELECT full_name, short_name, tax_id, base_accounting_fee, bank_account, id FROM companies WHERE is_active=0 ORDER BY full_name ASC')
        for f, s, t, fee, b, cid in self.db.cursor.fetchall():
            # 精進點：月記帳費去除小數點 .0
            self.inactive_company_tree.insert('', tk.END, values=("", f, s, t, int(round(fee)) if fee else 0, b, cid))

    def refresh_treeview(self, month):
        year = self.year_var.get()
        tree = self.tabs[month]
        for item in tree.get_children(): tree.delete(item)
            
        # 每次刷新期別時，先確認底冊是否已自動初始化
        self.check_and_initialize_billing(year, month)
        
        # 請款明細左側一律以公司編號遞增排序 (多選出 r.is_billed)
        self.db.cursor.execute('''
            SELECT c.company_code, r.is_paid, r.is_billed, c.full_name, c.bank_account, r.business_tax_amount, r.accounting_fee, r.note, r.id
            FROM billing_records r JOIN companies c ON r.company_internal_id = c.id
            WHERE r.year = ? AND r.month_period = ? AND c.is_active = 1
            ORDER BY c.company_code ASC
        ''', (year, month))
        
        for code, is_paid, is_billed, name, bank, tax, fee, note, rid in self.db.cursor.fetchall():
            paid_icon = '✅' if is_paid else '☐'
            billed_icon = '📨' if is_billed else '☐'
            
            self.db.cursor.execute("SELECT SUM(item_amount) FROM billing_special_items WHERE billing_record_id = ?", (rid,))
            sp_res = self.db.cursor.fetchone()
            sp_sum = sp_res[0] if sp_res[0] else 0.0
            
            # 確保運算變數安全非空
            tax_val = tax if tax is not None else 0.0
            fee_val = fee if fee is not None else 0.0
            total_amt = tax_val + fee_val + sp_sum
            
            # 精進點：金額全部轉換為無小數點整數，且格式化為帶有千分位逗號的字串 (僅用於 Treeview 顯示)
            tax_str = f"{int(round(tax_val)):,}"
            fee_str = f"{int(round(fee_val)):,}"
            sp_sum_str = f"{int(round(sp_sum)):,}"
            total_amt_str = f"{int(round(total_amt)):,}"
            
            tree.insert('', tk.END, text=code, values=(billed_icon, paid_icon, name, bank if bank else "", tax_str, fee_str, sp_sum_str, total_amt_str, note if note else "", rid))

   # ================= 產生請款單並開啟 Excel =================
    def generate_billing_excel(self):
        if not self.selected_record_id:
            messagebox.showwarning("警告", "請先選擇一筆公司的請款紀錄！")
            return
            
        template_file = "範本.xlsx"
        if not os.path.exists(template_file):
            messagebox.showerror("錯誤", f"找不到 {template_file}，請確認該檔案與程式放在同一個資料夾。")
            return

        try:
            # 1. 取得需要填入的資料
            # 從資料庫取得公司完整名稱
            self.db.cursor.execute('''
                SELECT c.full_name
                FROM billing_records r JOIN companies c ON r.company_internal_id = c.id WHERE r.id = ?
            ''', (self.selected_record_id,))
            res = self.db.cursor.fetchone()
            if not res: return
            company_name = res[0]

            # 處理月份字串 (將 "01-02月" 轉換為 "01月-02月")
            raw_month = self.current_month_var.get()
            months_split = raw_month.replace('月', '').split('-')
            if len(months_split) == 2:
                month_str_tax = f"{months_split[0]}月-{months_split[1]}月營業稅"
                month_str_fee = f"{months_split[0]}月-{months_split[1]}月記帳費"
            else:
                month_str_tax = f"{raw_month}營業稅"
                month_str_fee = f"{raw_month}記帳費"

            # 取得金額與備註
            try: tax_amt = float(self.entry_tax.get().strip())
            except: tax_amt = 0.0
            try: fee_amt = float(self.entry_accounting_fee.get().strip())
            except: fee_amt = 0.0
            note_str = self.entry_note.get().strip()

            # 取得動態特別項目
            sp_items = []
            for item in self.special_items_tree.get_children():
                vals = self.special_items_tree.item(item)['values']
                try:
                    # 強制轉為浮點數，解決 Excel 出現「Number stored as text」綠色警告
                    amt = float(vals[1])
                except ValueError:
                    amt = 0.0
                sp_items.append((vals[0], amt)) # (名稱, 金額)

            # 2. 載入 Excel 範本並寫入資料 (保留格式)
            wb = openpyxl.load_workbook(template_file)
            ws = wb.active # 預設操作第一個分頁
            
            # 辨識當營業稅金額是0的時候，填入"留抵"
            if tax_amt == 0.0:
                tax_amt = "留抵"

            # 寫入固定欄位
            ws['B2'].value = company_name
            ws['B4'].value = month_str_tax
            ws['B5'].value = month_str_fee
            ws['C4'].value = tax_amt
            ws['C5'].value = fee_amt
            ws['D9'].value = note_str

            # 先清空動態特別項目的舊資料 (避免覆蓋時殘留上一次的紀錄)
            for r in range(6, 10):
                ws[f'B{r}'].value = ""
                ws[f'C{r}'].value = ""

            # 依序填入動態項目 (設定上限最多 4 個，寫入 B6~B9, C6~C9)
            start_row = 6
            for i, (name, amt) in enumerate(sp_items[:4]):
                ws[f'B{start_row + i}'].value = name
                ws[f'C{start_row + i}'].value = amt

            # 直接覆蓋儲存原範本檔案
            wb.save(template_file)
            print(f"📄 已成功產生並覆蓋 {template_file}，正在為您開啟...")

            # 3. 呼叫系統預設程式開啟 Excel 視窗
            if os.name == 'nt': # Windows 系統
                os.startfile(template_file)
            else: # macOS / Linux 相容處理
                opener = "open" if sys.platform == "darwin" else "xdg-open"
                import subprocess
                subprocess.call([opener, template_file])

        except PermissionError:
            print("❌ 錯誤：無法覆蓋檔案。請確認「範本.xlsx」是否已經在 Excel 中開啟？")
            messagebox.showerror("檔案被鎖定", "無法覆蓋「範本.xlsx」！\n請確認該 Excel 檔案是否已經被開啟，請先關閉視窗後再試一次。")
        except Exception as e:
            print(f"❌ 產生請款單發生錯誤: {str(e)}")
            messagebox.showerror("錯誤", f"產生請款單時發生錯誤:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = TaxBillingApp(root)
    root.mainloop()
